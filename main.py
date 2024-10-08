import numpy as np
import splunklib.client as client
from dotenv import load_dotenv
import os
import xml.etree.ElementTree as ET
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import subprocess
import ipaddress

filePath = 'checkpoint.xlsx'


def connect_to_splunk(username, password, host='localhost', port='8089', owner='admin', app='search', sharing='user'):
    try:
        service = client.connect(host=host, port=port, username=username, password=password, owner=owner, app=app,
                                 sharing=sharing)
        if service:
            print("Splunk service created successfully")
            print("------------------------------------")
        return service
    except Exception as e:
        print(e)


def output_excel(jobResult):
    xmlData = str(jobResult)
    # 解析 XML 資料
    root = ET.fromstring(xmlData)

    data = []
    for result in root.findall('result'):
        src_ip = None
        dest_ips = []
        dest_ports = []
        counts = []

        for field in result.findall('field'):
            # 欄位名稱
            key = field.get('k')

            # 欄位值
            if key == 'src_ip':
                src_ip = field.find('value/text').text
            elif key == 'dest_ip':
                for value in field.findall('value'):
                    text_element = value.find('text')
                    if text_element is not None:
                        dest_ips.append(text_element.text)
            elif key == 'dest_port':
                for value in field.findall('value'):
                    text_element = value.find('text')
                    if text_element is not None:
                        dest_ports.append(text_element.text)
            elif key == 'count':
                for value in field.findall('value'):
                    text_element = value.find('text')
                    if text_element is not None:
                        counts.append(text_element.text)

        for dest_ip, dest_port, count in zip(dest_ips, dest_ports, counts):
            data.append([src_ip, dest_ip, dest_port, count])

    df = pd.DataFrame(data, columns=['src_ip', 'dest_ip', 'dest_port', 'count'])
    # 去除 src_ip 重複儲存格
    df['src_ip'] = df['src_ip'].where(~df['src_ip'].duplicated())
    df.to_excel(filePath, index=False)


def fetch_daily_report(splunk_service, search_string, payload={}):
    try:
        job = splunk_service.jobs.create(search_string, **payload)
        while True:
            while not job.is_ready():
                pass
            if job["isDone"] == "1":
                break

        output_excel(job.results())
    except Exception as e:
        print(e)


def is_ip_matching(ip, domain_name):
    # 取出域名的第一段
    domain_part = domain_name.split('.')[0]

    # 將 ip 轉成域名的格式，例如 10.10.10.10 轉成 10-10-10-10
    ip_part = ip.replace('.', '-')

    return ip_part == domain_part


def nslookup(ip):
    try:
        result = subprocess.run(['nslookup', ip], capture_output=True, text=True, check=True)

        # 解析 nslookup 输出
        output = result.stdout
        lines = output.splitlines()

        # 提取域名
        for line in lines:
            if '名稱:' in line:
                domain_name = line.split(':')[1].strip()
                return domain_name

        return "Domain name not found."

    except subprocess.CalledProcessError as e:
        return f"Error occurred: {e}"


def isPrivate(ip):
    try:
        ip = ipaddress.ip_address(ip)
    except ValueError:
        return False

    return ip.is_private


def dnsSearch(splunk_service, df, file_path, sheet_name):
    def get_fqdn(ip):
        try:
            if isPrivate(ip):
                return "Private IP"

            search_query = f"search index=infoblox {ip} | head 1 | table dns_answer_name"
            payload = {"exec_mode": "normal", "earliest_time": "-1d@d", "latest_time": "@d"}
            job = splunk_service.jobs.create(search_query, **payload)

            while True:
                while not job.is_ready():
                    pass
                if job["isDone"] == "1":
                    break

            xml_result = str(job.results())
            root = ET.fromstring(xml_result)
            FQDNs = ""
            fqdn_set = set()

            for result in root.findall('result'):
                for field in result.findall('field'):
                    # 欄位名稱
                    key = field.get('k')
                    # 欄位值
                    if key == 'dns_answer_name':
                        for value in field.findall('value'):
                            text_element = value.find('text')
                            if text_element is not None:
                                fqdn = text_element.text
                                if fqdn not in fqdn_set:
                                    fqdn_set.add(fqdn)
                                    FQDNs += fqdn + "\n"

            if FQDNs == "":
                FQDNs += nslookup(ip)

            if is_ip_matching(ip, FQDNs):
                print(f"IP = {ip}")
                print(f"Domain name = {FQDNs}")
                return ""

            print(ip)
            print(FQDNs + "\n")
            return FQDNs
        except Exception as e:
            return 'dnsSearch() Failed.'

    # 對每組 dest_ip 進行 DNS 查詢
    df['fqdn'] = df['dest_ip'].apply(get_fqdn)
    df.to_excel(file_path, sheet_name=sheet_name, index=False)

    '''
    # 美化表格 : 自動換行、置中對齊等等
    wb = load_workbook(file_path)
    ws = wb.active

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True)

    wb.save(file_path)
    '''


def main():
    try:
        load_dotenv()
        Username = os.getenv("username")
        Password = os.getenv("password")
        Host = os.getenv("host")

        splunk_service = connect_to_splunk(username=Username, password=Password, host=Host)

        # Daily Report 的語法
        index = 'checkpoint'
        search_query = (
            f"search index={index} action=Drop src_ip IN (10.*,172.16.*,172.18.*,172.19.*,172.17.*,192.168.*) dest!=10.*"
            "| eval dest_port=if(dest_port>1025,\"High Port(1025-65535)\",dest_port) "
            "| stats count by src_ip dest_port dest_ip "
            "| sort - count "
            "| stats list(dest_port) as dest_port list(dest_ip) as dest_ip list(count) as count by src_ip "
            "| sort - count "
            "| table src_ip dest_ip dest_port count "
            "| head 50")

        payload = {"exec_mode": "normal", "earliest_time": "-1d@d", "latest_time": "@d"}
        fetch_daily_report(splunk_service, search_query, payload)

        sheet_name = 'Sheet1'
        df = pd.read_excel(filePath, sheet_name=sheet_name)

        # 在 dest_ip 欄位的右邊增加一個 fqdn 欄位
        insert_index = 2
        columnName = 'fqdn'
        columnData = [np.nan] * len(df)  # 初始化為空值，長度與 DataFrame 行數匹配
        df.insert(insert_index, columnName, columnData)
        # df.to_excel(filePath, sheet_name=sheet_name, index=False)

        # 對每個 dest_ip 查詢 dns query
        dnsSearch(splunk_service, df, filePath, sheet_name)
        print(f"Excel 文件已生成：{filePath}")

    except Exception as e:
        print(e)


if __name__ == "__main__":
    main()